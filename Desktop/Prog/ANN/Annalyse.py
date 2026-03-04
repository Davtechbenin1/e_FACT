#Coding:utf-8
"""
	Surface de gestion des analyses de ventes, de recette et d'arrivage
"""
from lib.davbuild import *
from General_surf import *
import copy
from ..TRES.Trs_surf import Histo_Gene

class th_recycle_item(MyItem):
	def __init__(self,mother, **kwargs):
		#kwargs['bg_color'] = mother.sc.sep
		super().__init__(mother,**kwargs)
		self.height = dp(40)*7
		self._data_index = -1
		#self.spacing = dp(1)
		#self.padding = [dp(1),0,dp(1),dp(1)]

	@Cache_error
	def Foreign_surf(self,data_dict):
		self.clear_widgets()
		bo = box(self,size_hint = (1,1),orientation = 'horizontal',
			bg_color = self.sc.sep,padding = [dp(1),0,dp(1),dp(1)],
			spacing = dp(1))
		self.add_surf(bo)
		entetes = data_dict.get("entetes")
		H_list = list()
		ent_l = 3
		if 'articles' in entetes:
			ent_l+=3
			H_list.append((dp(40)+dp(1))*len(data_dict.get('articles')))
		if 'sous familles' in entetes:
			ent_l+=3
			H_list.append((dp(40)+dp(1))*len(data_dict.get('sous familles')))
		if 'familles' in entetes:
			ent_l+=3
			H_list.append((dp(40)+dp(1))*len(data_dict.get('familles')))
		#self.height = max(H_list)+dp(40)

		w = 1/ent_l
		if data_dict.get('articles'):
			for ent in entetes:
				info = data_dict.get(ent)
				if isinstance(info,dict):
					H = len(info)*dp(40)
					bbb = stack(self,size_hint = (1,None),
						bg_color = self.sc.aff_col3,
						height = H)
					for name,name_d in info.items():
						name_d['Nom'] = name
						bc = box(self,size_hint = (1,None),
							bg_color = self.sc.sep,height = dp(40),
							orientation = 'horizontal',
							)#padding = dp(1))
						for k in ['Nom',"Qté","Montant"]:
							v = name_d.get(k)
							bc.add_text(self.format_val(v,3),
								text_color = self.sc.text_col1,
								padding_left = dp(10),
								bg_color = self.sc.aff_col3)
						bbb.add_surf(bc)
					sc = scroll(self,size_hint = (w*3,1),
						bg_color = self.sc.aff_col3)
					sc.add_surf(bbb)
					bo.add_surf(sc)

				else:
					bc = box(self,size_hint = (w,1),bg_color = self.sc.sep,
						#spadding_right = dp(1)
						)
					bc.add_text(self.format_val(info,3),
						text_color = self.sc.text_col1,
						padding_left = dp(10),valign = 'top',
						bg_color = self.sc.aff_col3,padding_top = dp(5))
					bo.add_surf(bc)

class th_recycle_item1(MyItem):
	def __init__(self,mother, **kwargs):
		super().__init__(mother,**kwargs)
		self.height = dp(40)*7
		self._data_index = -1

	@Cache_error
	def Foreign_surf(self,data_dict):
		self.clear_widgets()
		bo = box(self,size_hint = (1,1),orientation = 'horizontal',
			padding = [dp(1),0,dp(1),dp(1)],spacing = dp(1),bg_color = self.sc.sep)
		self.add_surf(bo)
		entetes = data_dict.get("entetes")
		H_list = list()
		ent_l = 6
		w = 1/ent_l
		if data_dict.get('clients'):
			for ent in entetes:
				info = data_dict.get(ent)
				if isinstance(info,dict):
					H = len(info)*dp(40)
					bbb = stack(self,size_hint = (1,None),
						bg_color = self.sc.aff_col3,
						height = H)
					for name,name_d in info.items():
						name_d['Nom'] = name
						bc = box(self,size_hint = (1,None),
							bg_color = self.sc.sep,height = dp(40),
							orientation = 'horizontal',
							)#padding = dp(1))
						for k in ['Nom',"nombre de recette","Montant"]:
							v = name_d.get(k)
							bc.add_text(self.format_val(v,3),
								text_color = self.sc.text_col1,
								padding_left = dp(10),
								bg_color = self.sc.aff_col3)
						bbb.add_surf(bc)
					sc = scroll(self,size_hint = (w*3,1),
						bg_color = self.sc.aff_col3)
					sc.add_surf(bbb)
					bo.add_surf(sc)

				else:
					bc = box(self,size_hint = (w,1),bg_color = self.sc.sep,
						)
					bc.add_text(self.format_val(info,3),
						text_color = self.sc.text_col1,
						padding_left = dp(10),valign = 'top',
						bg_color = self.sc.aff_col3,padding_top = dp(5))
					bo.add_surf(bc)

class Annalyse(menu_surf_V_maquette):
	def Get_menu_infos(self):
		self.day1 = self.sc._get_yerterday(30)
		self.wid_dict = {
			"Ventes":Ventes,
			"Recettes":recette,
			"Achats":Arrivage,
			"Historiques Comptable global":Histo_Gene,
		}
		self.icon_dict = {
			"Recettes":"currency-usd",
			"Ventes":"cart-outline",
			"Achats":"truck",
			"Historiques Comptable global":'chart-line',
		}

class Ventes(stack):
	def __init__(self,mother,**kwargs):
		kwargs['bg_color'] = mother.sc.aff_col1
		kwargs['radius'] = dp(10)
		stack.__init__(self,mother,**kwargs)
		self.padding = dp(10)
		self.spacing = dp(5)

	@Cache_error
	def initialisation(self):
		self.day1 = self.mother.day1
		self.day2 = self.mother.day2
		self.art_info = 'Ajouter'
		self.s_fam_info = str()
		self.fam_info = str()
		self.size_pos()

	def size_pos(self):
		h = .04
		self.add_text("Statistique générale des ventes",halign ="center",
			text_color = self.sc.text_col1,size_hint = (.3,h),font_size = "17sp",
			underline = True)
		self.add_text("Quantité totale :",size_hint = (.08,h),
			text_color = self.sc.text_col1)
		self.quatite_total_surf = self.add_text(str(),
			text_color = self.sc.green,
			size_hint = (.12,h),font_size = "18sp")
		self.add_text("Montant total :",size_hint = (.08,h),
			text_color = self.sc.text_col1)
		self.montant_total_surf = self.add_text(str(),
			text_color = self.sc.green,
			size_hint = (.12,h),font_size = "18sp")
		self.add_surf(Periode_set(self,exc_fonc = self.Fill,
			size_hint = (.25,h),info_w = .15))
		self.add_icon_but(icon = 'printer',text_color = self.sc.black,
			size_hint = (.05,h),on_press = self.Impression)
		dic = {
			"Ajouter les articles ?":(self.art_info,['Ajouter'],
				self.add_th_art),
			"Ajouter les sous familles ?":(self.s_fam_info,['Ajouter'],
				self.add_th_s_fam),
			"Ajouter les familles ?":(self.fam_info,['Ajouter'],
				self.add_th_fam)
		}
		for k,tup in dic.items():
			txt,lis,fonc = tup
			self.add_text(k,text_color = self.sc.text_col1,
				size_hint = (.15,h))
			self.add_surf(liste_set(self,txt,lis,mother_fonc = fonc,
				size_hint = (.18,h),mult = 1))
		self.Tabe = box(self,size_hint = (1,.91),
			bg_color = self.sc.aff_col3)
		self.add_surf(self.Tabe)
		self.Fill()

	def Fill(self,*args):
		Clock.schedule_once(self._Fill)
		#self._Fill()

	@Cache_error
	def _Fill(self,*args):
		self.quatite_total = float()
		self.montant_total = float()

		self.Tabe.clear_widgets()
		date_list = self.get_date_list(self.day1,self.day2)
		self.mother.day1 = self.day1
		self.mother.day2 = self.day2
		entetes = ["date",'quantité totale',"montant total"]
		ent_surf = box(self,orientation = 'horizontal',
			height = dp(70),size_hint = (1,None),
			bg_color = self.sc.sep,padding = dp(1),spacing = dp(1))
		ent_l = 3
		if self.art_info:
			entetes.append("articles")
			ent_l += 3

		if self.s_fam_info:
			entetes.append("sous familles")
			ent_l += 3

		if self.fam_info:
			entetes.append("familles")
			ent_l += 3

		w = 1/ent_l
		for u in entetes:
			if u in ["date",'quantité totale',"montant total"]:
				ent_surf.add_text(u,size_hint = (w,1),padding_left = dp(10),
					text_color = self.sc.text_col1,bg_color = self.sc.aff_col1)
			else:
				th_st = box(self,size_hint = (w*3,1),spacing = dp(1),)
				th_st.add_text(u,text_color = self.sc.text_col1,
					halign = 'center',bg_color = self.sc.aff_col1)
				b = box(self,orientation = "horizontal")
				for t in ['Nom',"Qté","Montant"]:
					b.add_text(t,bg_color = self.sc.aff_col1,halign = 'left',
						text_color = self.sc.text_col1,padding_left = dp(10))
				th_st.add_surf(b)
				ent_surf.add_surf(th_st)
		self.Tabe.add_surf(ent_surf)

		data = list()
		for dic in self.set_infos(date_list):
			dic['entetes'] = entetes
			self.montant_total += dic.get('montant total')
			self.quatite_total += dic.get("quantité totale")
			data.append(dic)
		self.imp_data = data
		data.sort(key = itemgetter("DATE"),reverse = True)
		scr = vscroll(self)
		scr.set_viewclass(th_recycle_item)
		scr.set_data(data)
		self.Tabe.add_surf(scr)
		self.Up_entet()

	@Cache_error
	def set_infos(self, date_liste):
		all_dic = dict()
		template = {
			"articles": dict(),
			"familles": dict(),
			"sous familles": dict(),
			"montant total": 0.0,
			"quantité totale": 0.0
		}

		cache_articles = {}

		for th_date in date_liste:
			cmd_infos = self.sc.DB.Get_all_commandes(th_date)
			for num in cmd_infos.keys():
				cmd_dic = self.sc.DB.Get_this_cmd(num)
				if cmd_dic.get('status de la commande') != "Livrée":
					continue

				date = cmd_dic.get("date d'émission")
				this_date_d = all_dic.get(date, copy.deepcopy(template))
				this_date_d["date"] = date
				this_date_d['DATE'] = datetime.strptime(date,self.date_format)

				for art in cmd_dic.get('articles', []):
					art_name = art.get('Désignation')
					mont_ht  = art.get('montant HT')

					# Cache article
					art_d = cache_articles.get(art_name)
					if not art_d:
						art_d = self.sc.DB.Get_this_article(art_name)
						cache_articles[art_name] = art_d
					if art_d:
						# Quantité
						qte = art.get('qté')
						uni = art.get('unité')
						if uni:
							uni_qte = float(art_d.get('nbre_par_qté'))
							qte += round(uni/uni_qte, 4)
						art["qté"] = qte

						# Totaux
						this_date_d["quantité totale"] += qte
						this_date_d["montant total"] += mont_ht

						# Hiérarchie
						s_fam = art_d.get('sous famille')
						fam   = art_d.get('famille')
						my_art= art.get('Désignation').replace("_"," ")

						this_date_d["sous familles"] = self.set_part(this_date_d,"sous familles",s_fam,art)
						this_date_d["familles"] = self.set_part(this_date_d,"familles",fam,art)
						this_date_d["articles"] = self.set_part(this_date_d,"articles",my_art,art)

				all_dic[date] = this_date_d

		return all_dic.values()

	def set_part(self,this_date_d,part,part_txt,art):
		part_d = this_date_d.get(part)
		my_art = part_txt#art_d.get('Désignation').replace("_"," ")
		if my_art:
			my_art_d = part_d.get(my_art,{"Qté":float(),"Montant":float()})
			my_art_d["Qté"] += art.get("qté")
			my_art_d["Montant"] += art.get('montant HT')
			part_d[my_art] = my_art_d
		return part_d

	def Up_entet(self):
		self.montant_total_surf.text = self.format_val(self.montant_total)
		self.quatite_total_surf.text = self.format_val(self.quatite_total)

# Gestion des actions des méthodes
	@Cache_error
	def Impression(self,wid):
		fic = export_data_to_excel(self,self.imp_data,
			header_message = f'Détails des ventes du {self.day1} au {self.day2}',
			)
		self.open_link(fic)

	def add_th_art(self,info):
		self.art_info = info
		self.Fill()
	
	def add_th_s_fam(self,info):
		self.s_fam_info = info
		self.Fill()

	def add_th_fam(self,info):
		self.fam_info = info
		self.Fill()

class Arrivage(Ventes):
	def size_pos(self):
		h = .04
		self.add_text("Statistique générale des arrivages",halign ="center",
			text_color = self.sc.text_col1,size_hint = (.3,h),font_size = "17sp",
			underline = True)
		self.add_text("Quantité totale :",size_hint = (.08,h),
			text_color = self.sc.text_col1)
		self.quatite_total_surf = self.add_text(str(),
			text_color = self.sc.green,
			size_hint = (.12,h),font_size = "18sp")
		self.add_text("Montant total :",size_hint = (.08,h),
			text_color = self.sc.text_col1)
		self.montant_total_surf = self.add_text(str(),
			text_color = self.sc.green,
			size_hint = (.12,h),font_size = "18sp")
		self.add_surf(Periode_set(self,exc_fonc = self.Fill,
			size_hint = (.25,h),info_w = .2))
		self.add_icon_but(icon = 'printer',text_color = self.sc.black,
			size_hint = (.05,h),on_press = self.Impression)
		
		dic = {
			"Ajouter les articles ?":(self.art_info,['Ajouter'],
				self.add_th_art),
			"Ajouter les sous familles ?":(self.s_fam_info,['Ajouter'],
				self.add_th_s_fam),
			"Ajouter les familles ?":(self.fam_info,['Ajouter'],
				self.add_th_fam)
		}
		for k,tup in dic.items():
			txt,lis,fonc = tup
			self.add_text(k,text_color = self.sc.text_col1,
				size_hint = (.15,h))
			self.add_surf(liste_set(self,txt,lis,mother_fonc = fonc,
				size_hint = (.18,h),mult = 1))
		self.Tabe = box(self,size_hint = (1,.91),
			bg_color = self.sc.aff_col3)
		self.add_surf(self.Tabe)
		self.Fill()

	@Cache_error
	def _Fill(self,*args):
		self.quatite_total = float()
		self.montant_total = float()

		self.Tabe.clear_widgets()
		date_list = self.get_date_list(self.day1,self.day2)
		entetes = ["date",'quantité totale',"montant total"]
		ent_surf = box(self,orientation = 'horizontal',
			height = dp(70),size_hint = (1,None),
			bg_color = self.sc.sep,padding = dp(1),spacing = dp(1))
		ent_l = 3
		if self.art_info:
			entetes.append("articles")
			ent_l += 3

		if self.s_fam_info:
			entetes.append("sous familles")
			ent_l += 3

		if self.fam_info:
			entetes.append("familles")
			ent_l += 3

		w = 1/ent_l
		for u in entetes:
			if u in ["date",'quantité totale',"montant total"]:
				ent_surf.add_text(u,size_hint = (w,1),padding_left = dp(10),
					text_color = self.sc.text_col1,bg_color = self.sc.aff_col1)
			else:
				th_st = box(self,size_hint = (w*3,1),spacing = dp(1),)
				th_st.add_text(u,text_color = self.sc.text_col1,
					halign = 'center',bg_color = self.sc.aff_col1)
				b = box(self,spacing = dp(1),orientation = "horizontal")
				for t in ['Nom',"Qté","Montant"]:
					b.add_text(t,bg_color = self.sc.aff_col1,halign = 'center',
						text_color = self.sc.text_col1)
				th_st.add_surf(b)
				ent_surf.add_surf(th_st)
		self.Tabe.add_surf(ent_surf)

		data = list()
		for dic in self.set_infos(date_list):
			dic['entetes'] = entetes
			self.montant_total += dic.get('montant total')
			self.quatite_total += dic.get("quantité totale")
			data.append(dic)
		self.imp_data = data
		data.sort(key = itemgetter("DATE"),reverse = True)
		scr = vscroll(self)
		scr.set_viewclass(th_recycle_item)
		scr.set_data(data)
		self.Tabe.add_surf(scr)
		self.Up_entet()

	def set_infos(self, date_liste):
		all_dic = dict()
		template = {
			"articles": dict(),
			"familles": dict(),
			"sous familles": dict(),
			"montant total": 0.0,
			"quantité totale": 0.0
		}

		cache_articles = {}

		
		cmd_infos = self.sc.DB.Get_arrivage_from(date_liste)
		for cmd_dic in cmd_infos:
			date = cmd_dic.get("date")
			this_date_d = all_dic.get(date, copy.deepcopy(template))
			this_date_d["date"] = date
			this_date_d['DATE'] = datetime.strptime(date,self.date_format)

			for art in cmd_dic.get('articles', []):
				art_name = art.get('nom')
				mont_ht  = art.get('montant_HT')

				# Cache article
				art_d = cache_articles.get(art_name)
				if not art_d:
					art_d = self.sc.DB.Get_this_article(art_name)
					cache_articles[art_name] = art_d
				if art_d:
				# Quantité
					qte = art.get('qté')
					uni = art.get('unité')
					if uni:
						uni_qte = float(art_d.get('nbre_par_qté'))
						qte += round(uni/uni_qte, 4)
					art["qté"] = qte

					# Totaux
					this_date_d["quantité totale"] += qte
					this_date_d["montant total"] += mont_ht

					# Hiérarchie
					s_fam = art_d.get('sous famille')
					fam   = art_d.get('famille')
					my_art= art.get('Désignation').replace("_"," ")

					this_date_d["sous familles"] = self.set_part(this_date_d,"sous familles",s_fam,art)
					this_date_d["familles"] = self.set_part(this_date_d,"familles",fam,art)
					this_date_d["articles"] = self.set_part(this_date_d,"articles",my_art,art)


			all_dic[date] = this_date_d

		return all_dic.values()

	def set_part(self,this_date_d,part,part_txt,art):
		part_d = this_date_d.get(part)
		my_art = part_txt#art_d.get('Désignation').replace("_"," ")
		if my_art:
			my_art_d = part_d.get(my_art,{"Qté":float(),"Montant":float()})
			my_art_d["Qté"] += art.get("qté")
			my_art_d["Montant"] += art.get('montant_HT')
			part_d[my_art] = my_art_d
		return part_d

	@Cache_error
	def Impression(self,wid):
		fic = export_data_to_excel(self,self.imp_data,
			header_message = f'Détails des approvisionnements du {self.day1} au {self.day2}',
			)
		self.open_link(fic)

class recette(Ventes):
	def size_pos(self):
		h = .04
		self.add_text("Statistique générale des Recettes",halign ="center",
			text_color = self.sc.text_col1,size_hint = (.3,h),font_size = "17sp",
			underline = True)
		self.add_text("Quantité totale :",size_hint = (.08,h),
			text_color = self.sc.text_col1)
		self.quatite_total_surf = self.add_text(str(),
			text_color = self.sc.green,
			size_hint = (.12,h),font_size = "18sp")
		self.add_text("Montant total :",size_hint = (.08,h),
			text_color = self.sc.text_col1)
		self.montant_total_surf = self.add_text(str(),
			text_color = self.sc.green,
			size_hint = (.12,h),font_size = "18sp")
		self.add_surf(Periode_set(self,exc_fonc = self.Fill,
			size_hint = (.25,h),info_w = .2))
		self.add_icon_but(icon = 'printer',text_color = self.sc.black,
			size_hint = (.05,h),on_press = self.Impression)
		
		self.Tabe = box(self,size_hint = (1,.955),
			bg_color = self.sc.aff_col3)
		self.add_surf(self.Tabe)
		self.Fill()

	@Cache_error
	def _Fill(self,*args):
		self.quatite_total = float()
		self.montant_total = float()

		self.Tabe.clear_widgets()
		date_list = self.get_date_list(self.day1,self.day2)
		entetes = ["date",'recette totale',"montant total",'clients']
		ent_surf = box(self,orientation = 'horizontal',
			height = dp(70),size_hint = (1,None),
			bg_color = self.sc.sep,padding = dp(1),spacing = dp(1))
		ent_l = 6

		w = 1/ent_l
		for u in entetes:
			if u in ["date",'recette totale',"montant total"]:
				ent_surf.add_text(u,size_hint = (w,1),padding_left = dp(10),
					text_color = self.sc.text_col1,bg_color = self.sc.aff_col1)
			else:
				th_st = box(self,size_hint = (w*3,1),spacing = dp(1),)
				th_st.add_text(u,text_color = self.sc.text_col1,
					halign = 'center',bg_color = self.sc.aff_col1)
				b = box(self,spacing = dp(1),orientation = "horizontal")
				for t in ['Nom',"nombre de recette","Montant"]:
					b.add_text(t,bg_color = self.sc.aff_col1,halign = 'center',
						text_color = self.sc.text_col1)
				th_st.add_surf(b)
				ent_surf.add_surf(th_st)
		self.Tabe.add_surf(ent_surf)

		data = list()
		for dic in self.set_infos(date_list):
			dic['entetes'] = entetes
			self.montant_total += dic.get('montant total')
			self.quatite_total += dic.get("recette totale")
			data.append(dic)
		self.imp_data = data
		data.sort(key = itemgetter("DATE"),reverse = True)
		scr = vscroll(self)
		scr.set_viewclass(th_recycle_item1)
		scr.set_data(data)
		self.Tabe.add_surf(scr)
		self.Up_entet()

	def set_infos(self, date_liste):
		all_dic = dict()
		template = {
			"clients": dict(),
			"montant total": 0.0,
			"recette totale": 0.0
		}

		cache_articles = {}

		all_paie_liste = self.sc.DB.Get_all_paiement_of(date_liste)
		for paie_dic in all_paie_liste:
			date = paie_dic.get("date de paiement")
			this_date_d = all_dic.get(date, copy.deepcopy(template))
			this_date_d["date"] = date
			this_date_d['DATE'] = datetime.strptime(date,self.date_format)

			this_date_d["montant total"]+=paie_dic.get('montant payé')
			this_date_d['recette totale'] += 1
			num_client = paie_dic.get('client')
			nom_client = self.sc.DB.Get_this_clt(num_client).get('nom',str())

			clt_d = this_date_d['clients'].get(nom_client,
				{"nombre de recette":int(),
				"Montant":float()})
			clt_d["nombre de recette"] += 1
			clt_d['Montant'] += paie_dic.get('montant payé')
			this_date_d["clients"][nom_client] = clt_d
			all_dic[date] = this_date_d
		return all_dic.values()

	@Cache_error
	def Impression(self,wid):
		sections = {"clients":['Nom',"nombre de recette","Montant"]}
		fic = export_data_to_excel(self,self.imp_data,
			header_message = f'Détails des recettes du {self.day1} au {self.day2}',
			sections = sections)
		self.open_link(fic)


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

@Cache_error
def export_data_to_excel(self,
	data,
	filename="rapport.xlsx",
	header_message="Rapport des ventes",
	sections=None
):
	if sections is None:
		sections = {
			"articles": ["Nom", "Qté", "Montant"],
			"sous familles": ["Nom", "Qté", "Montant"],
			"familles": ["Nom", "Qté", "Montant"]
		}

	wb = Workbook()
	ws = wb.active
	ws.title = "Rapport"

	# --- Styles ---
	default_font = Font(size=13)
	header_font = Font(bold=True, color="000000", size=13)
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	border = Border(
		left=Side(style="thin"), right=Side(style="thin"),
		top=Side(style="thin"), bottom=Side(style="thin")
	)
	header_fill = PatternFill(start_color="E8EBF7", end_color="E8EBF7", fill_type="solid")
	title_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
	total_fill = PatternFill(start_color="66CC66", end_color="66CC66", fill_type="solid")
	grand_total_fill = PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid")

	# --- 1. Titre ---
	start_col = 3
	total_cols = 1 + sum(len(v) for v in sections.values())
	end_col_letter = get_column_letter(start_col + total_cols - 1)

	ws.merge_cells(f"C1:{end_col_letter}1")
	ws["C1"] = header_message
	ws["C1"].font = Font(bold=True, size=16, color="000000")
	ws["C1"].alignment = center
	ws["C1"].fill = title_fill

	# Ligne de départ pour tableau
	row0 = 3

	# --- 2. En-têtes dynamiques ---
	col_index = start_col
	ws.cell(row=row0, column=col_index, value="Date")
	ws.merge_cells(start_row=row0, start_column=col_index, end_row=row0+1, end_column=col_index)
	col_index += 1

	nom_cols_indices = []  # pour se souvenir des colonnes "Nom"

	for section, subcols in sections.items():
		span = len(subcols)
		ws.cell(row=row0, column=col_index, value=section)
		ws.merge_cells(start_row=row0, start_column=col_index, end_row=row0, end_column=col_index+span-1)
		for i, colname in enumerate(subcols):
			ws.cell(row=row0+1, column=col_index+i, value=colname)
			if colname.lower() == "nom":
				nom_cols_indices.append(col_index+i)
		col_index += span

	# Style en-têtes
	for row in ws[f"{get_column_letter(start_col)}{row0}:{end_col_letter}{row0+1}"]:
		for cell in row:
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = center
			cell.border = border

	# --- 3. Largeur colonnes ---
	for col in range(start_col, start_col + total_cols):
		if col in nom_cols_indices:
			ws.column_dimensions[get_column_letter(col)].width = 30
		else:
			ws.column_dimensions[get_column_letter(col)].width = 18

	# --- 4. Remplissage des données ---
	current_row = row0 + 2
	mois_qte_total = 0
	mois_montant_total = 0

	for entry in data:
		date = entry.get("date", "")
		qte_tot = entry.get("quantité totale", 0)
		montant_tot = entry.get("montant total", 0)

		mois_qte_total += qte_tot
		mois_montant_total += montant_tot

		max_len = max(len(entry.get(sec, {})) for sec in sections)

		for i in range(max_len):
			row = current_row + i

			# Date
			if i == 0:
				cell_date = ws.cell(row=row, column=start_col, value=date)
				cell_date.font = Font(size=13, color="008000", bold=True)
				cell_date.alignment = center

			# Sections
			col_index = start_col + 1
			for sec, subcols in sections.items():
				sec_dict = entry.get(sec, {})
				keys = list(sec_dict.keys())
				if i < len(keys):
					name = keys[i]
					vals = sec_dict[name]
					for j, colname in enumerate(subcols):
						val = None
						if colname.lower() == "nom":
							val = name
						elif colname.lower() == "qté":
							val = self.format_val(vals.get("Qté", 0), 3)
						elif colname.lower() == "montant":
							val = self.format_val(vals.get("Montant", 0), 3)
						ws.cell(row=row, column=col_index+j, value=val)
				col_index += len(subcols)

		# Ligne total par entrée
		total_row = current_row + max_len
		ws.cell(row=total_row, column=start_col+1, value="TOTAL")

		qte_col = start_col + 2
		montant_col = start_col + 3
		ws.cell(row=total_row, column=qte_col, value=self.format_val(qte_tot, 3))
		ws.cell(row=total_row, column=montant_col, value=self.format_val(montant_tot, 3))

		# Style TOTAL
		for c in range(start_col, start_col+total_cols):
			cell = ws.cell(row=total_row, column=c)
			cell.font = Font(bold=True, size=13, color="FFFFFF")
			cell.fill = total_fill
			cell.alignment = center
			cell.border = Border(
				left=Side(style="thin"), right=Side(style="thin"),
				top=Side(style="thin"), bottom=Side(style="thin")
			)
		current_row = total_row + 2

	# --- 5. Total du mois ---
	ws.cell(row=current_row, column=start_col+1, value="TOTAL DU MOIS")
	ws.cell(row=current_row, column=qte_col, value=self.format_val(mois_qte_total, 3))
	ws.cell(row=current_row, column=montant_col, value=self.format_val(mois_montant_total, 3))

	# Style TOTAL DU MOIS
	for c in range(start_col, start_col+total_cols):
		cell = ws.cell(row=current_row, column=c)
		cell.font = Font(bold=True, size=14, color="FFFFFF")
		cell.fill = grand_total_fill
		cell.alignment = center
		cell.border = Border(
			left=Side(style="thin"), right=Side(style="thin"),
			top=Side(style="thin"), bottom=Side(style="thin")
		)

	# Police par défaut
	for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=start_col, max_col=start_col+total_cols-1):
		for cell in row:
			if cell.value is not None and not cell.font.bold:
				cell.font = default_font

	wb.save(filename)
	return filename

